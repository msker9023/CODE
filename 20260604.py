from  pathlib import Path
p=Path.cwd()
# context="CSV(Comma Separated Values)全称逗号分隔值文件是一种简单、通用的文件格式，被广泛的应用于应用程序（数据库、电子表格等）数据的导入和导出以及异构系统之间的数据交换。因为CSV是纯文本文件,不管是什么操作系统和编程语言都是可以处理纯文本的,而且很多编程语言中都提供了对读写CSV文件的支持,因此CSV格式在数据处理和数据科学中被广泛应用。" 
# Path(p/'context.txt').write_text(context,encoding='utf-8')
# for py in p.glob('*.txt'):
#     print(py)
# for item in p.iterdir():
#     print(item)
# import json

# my_dict = {
#     'name': '骆昊',
#     'age': 40,
#     'friends': ['王大锤', '白元芳'],
#     'cars': [
#         {'brand': 'BMW', 'max_speed': 240},
#         {'brand': 'Audi', 'max_speed': 280},
#         {'brand': 'Benz', 'max_speed': 280}
#     ]
# }
# with open(p/'data.json','w') as file:
#     json.dump(my_dict,file)
# with open(Path(p/'data.json'),'r') as file:
#     print(json.dumps(my_dict))
#     print(json.load(file))
    
# import csv
# import random
# from pathlib import Path
# p=Path.cwd()
# with open(p/'scores.csv','w') as file:
#     writer=csv.writer(file,delimiter='|')
#     writer.writerow(['姓名', '语文', '数学', '英语'])
#     names= ['关羽', '张飞', '赵云', '马超', '黄忠']
#     for name in names:
#         scores=[random.randrange(50,101) for _ in range(3)]
#         scores.insert(0, name)
#         writer.writerow(scores)
# with open(p/'scores.csv','r') as file:
#     reader=csv.reader(file,delimiter='|')
#     for data_list in reader:
#         print(reader.line_num,end='\t')
#         for elem in data_list:
#             print(elem,end='\t')
#         print()

import xlrd

# wb=xlrd.open_workbook('2022年股票数据.xls')
# sheetnames=wb.sheet_names()
# # print(sheetnames)
# sheet=wb.sheet_by_name(sheetnames[0])
# # print(sheet)
# #获取行数和列数
# print(sheet.nrows,sheet.ncols)
# for row in range(sheet.nrows):
#     for col in range(sheet.ncols):
#         value=sheet.cell(row,col).value
#         if row>0:
#             if col==0:
#                 value=xlrd.xldate_as_tuple(value,0)
#                 value=f'{value[0]}年{value[1]:>02d}月{value[2]:>02d}日'
#             else:
#                 value=f'{value:.2f}'
            
#         print(value,end='\t')
#     print()
# # 获取最后一个单元格的数据类型
# # 0 - 空值，1 - 字符串，2 - 数字，3 - 日期，4 - 布尔，5 - 错误
# last_cell_type=sheet.cell_type(sheet.nrows-1,sheet.ncols-1)
# last_cell=sheet.cell(1,sheet.ncols-1).value
# print(last_cell)
# print(last_cell_type)
# print(sheet.row_values(1))
# # 获取指定行指定列范围的数据（列表）
# # 第一个参数代表行索引，第二个和第三个参数代表列的开始（含）和结束（不含）索引
# print(sheet.row_slice(3,0,5))
import datetime
import openpyxl
wb=openpyxl.load_workbook('2022年股票数据.xlsx')
# print(wb.sheetnames)
sheet=wb.worksheets[0]
#dimension:尺寸,规模
# print(sheet.dimensions)
# print(sheet.max_row,sheet.max_column)
# print(sheet.cell(3,3).value)
# print(sheet['C3'].value)
# print(sheet['G255'].value)
# print(sheet['A2:C5'])
for row_ch in range(2,sheet.max_row+1):
    for col_ch in 'ABCDEFG':
        value=sheet[f'{col_ch}{row_ch}'].value
        if type(value)==datetime.datetime:
            print(value.strftime('%Y年%m月%d日'), end='\t')
        elif type(value)==int:
            print(f'{value:<4d}',end='\t')
        elif type(value)==float:
            print(f'{value:.4f}',end='\t')
        else:
            print(value,end='\t')
    print()


